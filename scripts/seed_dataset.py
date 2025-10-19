from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from sqlalchemy import delete
from sqlalchemy.orm import Session

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
    raise RuntimeError(
        "openpyxl is required to seed the dataset. Install it with 'pip install openpyxl'."
    ) from exc

from app.infrastructure.config import DatabaseConfig
from app.infrastructure.db import make_engine_and_session
from app.infrastructure.models import (
    Base,
    AcronymORM,
    DimensionORM,
    ExplanationORM,
    RatingScaleORM,
    ThemeLevelGuidanceORM,
    ThemeORM,
    TopicORM,
)

RatingColumn = tuple[int, str, str]


DIMENSION_IMAGE_MAP: dict[str, str] = {
    "Governance & Leadership": "governance-leadership.jpg",
    "Risk Assessment & Management": "risk-assessment-management.jpg",
    "BC & DR Planning": "bc-dr-planning.jpg",
    "Process & Dependency Mapping": "process-dependency-mapping.jpg",
    "IT & Cyber Resilience": "it-cyber-resilience.jpg",
    "Crisis Comms & Incident Mgmt": "crisis-comms-incident-mgmt.jpg",
    "Third-Party Resilience": "third-party-resilience.jpg",
    "Culture & Human Factors": "culture-human-factors.jpg",
    "Regulatory Compliance & Resolvability": "regulatory-compliance-resolvability.jpg",
}


def clean_text(value: object) -> str:
    if isinstance(value, str):
        return value.strip()
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_optional(value: object) -> str | None:
    text = clean_text(value)
    return text or None


def split_bullets(cell: object) -> list[str]:
    if not isinstance(cell, str):
        return []
    parts = re.split(r"[\n\r]+|\u2022", cell)
    cleaned = [p.strip(" \t-•") for p in parts if p and p.strip(" \t-•")]
    return cleaned


def detect_rating_columns(columns: Iterable[str]) -> list[RatingColumn]:
    pattern = re.compile(r"^(\d)\s+(.+)$")
    result: list[RatingColumn] = []
    for col in columns:
        match = pattern.match(col.strip())
        if match:
            level = int(match.group(1))
            label = match.group(2).strip()
            result.append((level, label, col))
    result.sort(key=lambda item: item[0])
    return result


@dataclass(slots=True)
class ExcelTable:
    name: str
    sheet: str
    ref: str
    dataframe: pd.DataFrame


class ExcelSeedSource:
    def __init__(self, tables: dict[str, ExcelTable]):
        self._tables = tables
        self._by_lower = {name.lower(): table for name, table in tables.items()}

    @classmethod
    def from_workbook(cls, excel_path: Path) -> "ExcelSeedSource":
        workbook = load_workbook(excel_path, data_only=True)
        tables: dict[str, ExcelTable] = {}
        try:
            for worksheet in workbook.worksheets:
                for table in worksheet.tables.values():
                    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                    raw_rows = [
                        list(row)
                        for row in worksheet.iter_rows(
                            min_row=min_row,
                            max_row=max_row,
                            min_col=min_col,
                            max_col=max_col,
                            values_only=True,
                        )
                    ]
                    if not raw_rows:
                        continue

                    headers: list[str] = []
                    for idx, cell in enumerate(raw_rows[0], start=1):
                        header = clean_text(cell)
                        if not header:
                            header = f"Column{idx}"
                        headers.append(header)

                    data_rows = raw_rows[1:]
                    dataframe = pd.DataFrame(data_rows, columns=headers)
                    tables[table.displayName] = ExcelTable(
                        name=table.displayName,
                        sheet=worksheet.title,
                        ref=table.ref,
                        dataframe=dataframe,
                    )
        finally:
            workbook.close()
        return cls(tables)

    def table_names(self) -> list[str]:
        return sorted(self._tables.keys())

    def registry(self) -> dict[str, dict[str, object]]:
        return {
            name: {
                "sheet": table.sheet,
                "ref": table.ref,
                "rows": int(table.dataframe.shape[0]),
                "columns": list(table.dataframe.columns),
            }
            for name, table in self._tables.items()
        }

    def require(self, name: str) -> ExcelTable:
        table = self._tables.get(name)
        if table:
            return table
        table = self._by_lower.get(name.lower())
        if table:
            return table
        available = ", ".join(self.table_names())
        raise KeyError(f"Table '{name}' not found in workbook. Available tables: {available}")

    def require_any(self, names: Iterable[str]) -> ExcelTable:
        for name in names:
            table = self._tables.get(name)
            if table:
                return table
            table = self._by_lower.get(name.lower())
            if table:
                return table
        available = ", ".join(self.table_names())
        raise KeyError(
            f"None of the tables {', '.join(names)} were found in workbook. Available tables: {available}"
        )

    def optional(self, name: str) -> ExcelTable | None:
        return self._tables.get(name) or self._by_lower.get(name.lower())


def extract_cmmi_definitions(table: pd.DataFrame) -> dict[int, dict[str, str | None]]:
    pattern = re.compile(r"^(\d)\s*(.+)$")
    mapping: dict[int, dict[str, str | None]] = {}
    for _, row in table.iterrows():
        raw_level = clean_text(row.get("Level"))
        if not raw_level:
            continue
        match = pattern.match(raw_level)
        if not match:
            continue
        level = int(match.group(1))
        label = match.group(2).strip()
        mapping[level] = {
            "label": label,
            "description": clean_optional(row.get("Definition")),
        }
    return mapping


def extract_dimension_descriptions(table: pd.DataFrame) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {}
    for _, row in table.iterrows():
        dimension = clean_text(row.get("Dimension"))
        if not dimension:
            continue
        description = clean_optional(row.get("Dimension_Description"))
        if dimension not in mapping or mapping[dimension] is None:
            mapping[dimension] = description
    return mapping


def extract_theme_descriptions(table: pd.DataFrame) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {}
    for _, row in table.iterrows():
        theme = clean_text(row.get("Theme"))
        if not theme:
            continue
        description = clean_optional(row.get("Theme_Description"))
        if theme not in mapping or mapping[theme] is None:
            mapping[theme] = description
    return mapping


def extract_topic_descriptions(table: pd.DataFrame) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {}
    for _, row in table.iterrows():
        topic = clean_text(row.get("Topic"))
        if not topic:
            continue
        description = clean_optional(row.get("Topic_Description"))
        if topic not in mapping or mapping[topic] is None:
            mapping[topic] = description
    return mapping


def extract_theme_generics(
    table: pd.DataFrame,
) -> tuple[dict[str, str | None], dict[str, dict[int, str]]]:
    categories: dict[str, str | None] = {}
    levels: dict[str, dict[int, str]] = {}
    for _, row in table.iterrows():
        theme = clean_text(row.get("Theme"))
        if not theme:
            continue
        categories[theme] = clean_optional(row.get("Category"))
        level_map: dict[int, str] = {}
        for level in range(1, 6):
            column = f"L{level} Generic"
            text = clean_text(row.get(column))
            if text:
                level_map[level] = text
        if level_map:
            levels[theme] = level_map
    return categories, levels


def sync_rating_scale(
    session: Session, rating_cols: list[RatingColumn], cmmi_info: dict[int, dict[str, str | None]]
) -> None:
    for level, fallback_label, _ in rating_cols:
        info = cmmi_info.get(level)
        label = info.get("label") if info else fallback_label
        description = info.get("description") if info else None
        existing = session.get(RatingScaleORM, level)
        if existing is None:
            session.add(RatingScaleORM(level=level, label=label, description=description))
        else:
            existing.label = label
            existing.description = description


def sync_theme_guidance(session: Session, theme: ThemeORM, guidance_levels: dict[int, str]) -> None:
    existing = {
        gl.level: gl
        for gl in session.query(ThemeLevelGuidanceORM)
        .filter(ThemeLevelGuidanceORM.theme_id == theme.id)
        .all()
    }
    for level, description in guidance_levels.items():
        entry = existing.get(level)
        if entry:
            entry.description = description
        else:
            session.add(
                ThemeLevelGuidanceORM(theme_id=theme.id, level=level, description=description)
            )
    stale_levels = set(existing) - set(guidance_levels)
    if stale_levels:
        session.execute(
            delete(ThemeLevelGuidanceORM).where(
                ThemeLevelGuidanceORM.theme_id == theme.id,
                ThemeLevelGuidanceORM.level.in_(list(stale_levels)),
            )
        )


def seed_acronyms(session: Session, table: pd.DataFrame) -> None:
    session.execute(delete(AcronymORM))
    for _, row in table.iterrows():
        acronym = clean_text(row.get("Acronym"))
        full_term = clean_text(row.get("Full term"))
        if not acronym or not full_term:
            continue
        meaning = clean_optional(row.get("Meaning / Why it matters in this framework"))
        session.add(
            AcronymORM(
                acronym=acronym,
                full_term=full_term,
                meaning=meaning,
            )
        )


def seed_from_excel(session: Session, excel_path: Path) -> ExcelSeedSource:
    source = ExcelSeedSource.from_workbook(excel_path)

    print("Discovered tables:")
    for name in source.table_names():
        entry = source.registry()[name]
        sheet = entry["sheet"]
        ref = entry["ref"]
        rows = entry["rows"]
        print(f" - {name}: sheet '{sheet}', range {ref}, rows={rows}")

    framework_df = source.require("Framework").dataframe
    cmmi_table = source.require_any(["CMM_definitions", "CMMI_definitions"]).dataframe
    dimension_desc_table = source.require("dimension_desc").dataframe
    theme_desc_table = source.require("theme_desc").dataframe
    topic_desc_table = source.require("topic_desc").dataframe
    theme_generic_table = source.require("theme_generic").dataframe

    cmmi_definitions = extract_cmmi_definitions(cmmi_table)
    dimension_descriptions = extract_dimension_descriptions(dimension_desc_table)
    theme_descriptions = extract_theme_descriptions(theme_desc_table)
    topic_descriptions = extract_topic_descriptions(topic_desc_table)
    theme_categories, theme_generic_levels = extract_theme_generics(theme_generic_table)

    rating_columns = detect_rating_columns(framework_df.columns)
    if not rating_columns:
        raise RuntimeError("No rating scale columns detected in the Framework table.")
    sync_rating_scale(session, rating_columns, cmmi_definitions)

    dimension_cache: dict[str, int] = {}
    theme_cache: dict[tuple[int, str], int] = {}
    processed_topics: set[int] = set()
    processed_themes: set[int] = set()

    for _, row in framework_df.iterrows():
        dimension_name = clean_text(row.get("Dimension"))
        theme_name = clean_text(row.get("Theme"))
        topic_name = clean_text(row.get("Topic"))

        if not (dimension_name and theme_name and topic_name):
            continue

        dimension_id = dimension_cache.get(dimension_name)
        if dimension_id is None:
            dimension = (
                session.query(DimensionORM).filter_by(name=dimension_name).one_or_none()
            )
            if dimension is None:
                dimension = DimensionORM(name=dimension_name)
                session.add(dimension)
                session.flush()
            dimension.description = dimension_descriptions.get(dimension_name)
            if dimension.image_filename is None:
                dimension.image_filename = DIMENSION_IMAGE_MAP.get(dimension_name)
            if dimension.image_alt is None:
                dimension.image_alt = dimension_name or None
            dimension_cache[dimension_name] = dimension.id
            dimension_id = dimension.id
        else:
            dimension = session.get(DimensionORM, dimension_id)
            if dimension:
                if dimension.description is None:
                    dimension.description = dimension_descriptions.get(dimension_name)
                if dimension.image_filename is None:
                    dimension.image_filename = DIMENSION_IMAGE_MAP.get(dimension_name)
                if dimension.image_alt is None:
                    dimension.image_alt = dimension_name or None

        theme_key = (dimension_id, theme_name)
        theme_id = theme_cache.get(theme_key)
        if theme_id is None:
            theme = (
                session.query(ThemeORM)
                .filter_by(dimension_id=dimension_id, name=theme_name)
                .one_or_none()
            )
            if theme is None:
                theme = ThemeORM(dimension_id=dimension_id, name=theme_name)
                session.add(theme)
                session.flush()
            theme.description = theme_descriptions.get(theme_name)
            theme.category = theme_categories.get(theme_name)
            theme_cache[theme_key] = theme.id
            theme_id = theme.id
        else:
            theme = session.get(ThemeORM, theme_id)
            if theme:
                if theme.description is None:
                    theme.description = theme_descriptions.get(theme_name)
                if theme.category is None:
                    theme.category = theme_categories.get(theme_name)

        if theme and theme.id not in processed_themes:
            guidance_levels = theme_generic_levels.get(theme_name)
            if guidance_levels:
                sync_theme_guidance(session, theme, guidance_levels)
            processed_themes.add(theme.id)

        topic = (
            session.query(TopicORM)
            .filter_by(theme_id=theme_id, name=topic_name)
            .one_or_none()
        )
        if topic is None:
            topic = TopicORM(theme_id=theme_id, name=topic_name)
            session.add(topic)
            session.flush()
        topic.description = topic_descriptions.get(topic_name)

        if topic.id not in processed_topics:
            session.execute(delete(ExplanationORM).where(ExplanationORM.topic_id == topic.id))
            processed_topics.add(topic.id)

        for level, _label, column in rating_columns:
            bullets = split_bullets(row.get(column))
            for bullet in bullets:
                session.add(ExplanationORM(topic_id=topic.id, level=level, text=bullet))

    acronyms_table = source.optional("Acronyms")
    if acronyms_table is not None:
        seed_acronyms(session, acronyms_table.dataframe)

    return source


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seed DB from Enhanced Operational Resilience Maturity workbook"
    )

    backend_default = os.environ.get("DB_BACKEND", "sqlite")
    sqlite_default = os.environ.get("DB_SQLITE_PATH") or os.environ.get("SQLITE_PATH", "./resilience.db")
    mysql_host_default = os.environ.get("DB_MYSQL_HOST") or os.environ.get("MYSQL_HOST", "localhost")
    mysql_port_default = int(
        os.environ.get("DB_MYSQL_PORT")
        or os.environ.get("MYSQL_PORT")
        or 3306
    )
    mysql_user_default = os.environ.get("DB_MYSQL_USER") or os.environ.get("MYSQL_USER", "root")
    mysql_password_default = os.environ.get("DB_MYSQL_PASSWORD") or os.environ.get("MYSQL_PASSWORD", "")
    mysql_database_default = os.environ.get("DB_MYSQL_DATABASE") or os.environ.get("MYSQL_DB", "resilience")

    parser.add_argument(
        "--backend", choices=["sqlite", "mysql"], default=backend_default
    )
    parser.add_argument("--sqlite-path", default=sqlite_default)
    parser.add_argument("--mysql-host", default=mysql_host_default)
    parser.add_argument("--mysql-port", type=int, default=mysql_port_default)
    parser.add_argument("--mysql-user", default=mysql_user_default)
    parser.add_argument("--mysql-password", default=mysql_password_default)
    parser.add_argument("--mysql-database", "--mysql-db", dest="mysql_database", default=mysql_database_default)
    parser.add_argument(
        "--excel-path",
        default="app/source_data/Enhanced_Operational_Resilience_Maturity_v6.xlsx",
    )
    args = parser.parse_args()

    cfg = DatabaseConfig(
        backend=args.backend,
        sqlite_path=args.sqlite_path,
        mysql_host=args.mysql_host,
        mysql_port=args.mysql_port,
        mysql_user=args.mysql_user,
        mysql_password=args.mysql_password,
        mysql_database=args.mysql_database,
    )
    url = cfg.get_connection_url()
    engine, SessionLocal = make_engine_and_session(url)

    Base.metadata.create_all(engine)

    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        print(f"ERROR: Excel file not found at {excel_path}", file=sys.stderr)
        sys.exit(1)

    with SessionLocal() as session:
        seed_from_excel(session, excel_path)
        session.commit()
    print("Seed completed.")


if __name__ == "__main__":
    main()
